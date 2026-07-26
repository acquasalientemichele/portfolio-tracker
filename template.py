"""
template.py — generazione in-memory del workbook Excel del portfolio tracker.

Perché un modulo invece di un file .xlsx statico nel repo:
- unica-fonte-di-verità: la struttura dei fogli (colonne, ordine) resta
  ancorata al codice e non va mai in drift con i validatori di portfolio.py
- l'app su Cloud non ha filesystem persistente: il template va servito come
  bytes generati al volo (st.download_button), non letto da disco
- la stessa infrastruttura genera i dati demo (modalità "prova senza compilare")

Il modulo è puro: nessun `import streamlit`. Restituisce sempre `bytes`,
pronti per st.download_button (template) o per st.session_state (demo).

Struttura del workbook (4 fogli, coerente con i loader esistenti):
  - transactions : date | ticker | isin | name | operation | quantity |
                   price | currency | fees | notes      (letto da pf.load_transactions)
  - settings     : base_currency / benchmark_ticker + blocco target allocation
                                                         (letto da pf.load_settings)
  - ter          : ticker | ter_annual | note            (letto da cst.load_costs, opzionale)
  - bollo_charges: date | amount | notes                 (letto da cst.load_costs, opzionale)

Dipendenze: openpyxl (già in requirements per add_costs_sheets).
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# STILE (allineato ad add_costs_sheets.py per coerenza visiva del workbook)
# --------------------------------------------------------------------------- #
NAVY = "1F4E78"
HDR_FILL = PatternFill("solid", start_color=NAVY)
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
NOTE_FONT = Font(italic=True, color="7F7F7F", name="Arial", size=10)
KEY_FONT = Font(bold=True, name="Arial", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

# Colonne del foglio transactions, nell'ordine atteso da pf.REQUIRED_COLS
# (+ 'notes', opzionale ma utile all'utente per annotare i PAC).
TX_HEADERS = ["date", "ticker", "isin", "name", "operation",
              "quantity", "price", "currency", "fees", "notes"]

# Anagrafica di default dei due ETF del PAC di riferimento. Usata sia per
# pre-compilare ter/settings nel template, sia per generare la demo.
_ETF_META = {
    "VWCE.DE": {
        "isin": "IE00BK5BQT80",
        "name": "Vanguard FTSE All-World (Acc)",
        "ter": 0.0022,
        "ter_note": "0,22% TER (Vanguard FTSE All-World)",
        "target": 0.80,
    },
    "VFEA.DE": {
        "isin": "IE00BK5BR733",
        "name": "Vanguard FTSE Emerging Markets (Acc)",
        "ter": 0.0022,
        "ter_note": "0,22% TER (Vanguard FTSE Emerging Markets)",
        "target": 0.20,
    },
}

# Anagrafica degli ETF usati nella DEMO. Tenuta separata da _ETF_META (che
# governa il template scaricabile) così i due possono evolvere in modo
# indipendente: modificando qui cambi solo il portafoglio di prova.
# Ticker Xetra: yfinance li quota, quindi la demo mostra prezzi reali.
_DEMO_META = {
    "VWCE.DE": {
        "isin": "IE00BK5BQT80",
        "name": "Vanguard FTSE All-World UCITS ETF (Acc)",
        "ter": 0.0022,
        "ter_note": "0,22% TER — azionario globale developed + emerging",
        "target": 0.66,
    },
    "SXR8.DE": {
        "isin": "IE00B5BMR087",
        "name": "iShares Core S&P 500 UCITS ETF (Acc)",
        "ter": 0.0007,
        "ter_note": "0,07% TER — large cap USA",
        "target": 0.17,
    },
    "EXSA.DE": {
        "isin": "DE0002635307",
        "name": "iShares STOXX Europe 600 UCITS ETF (Dist)",
        "ter": 0.0020,
        "ter_note": "0,20% TER — azionario Europa, a distribuzione",
        "target": 0.17,
    },
}

# --------------------------------------------------------------------------- #
# HELPER DI STILE
# --------------------------------------------------------------------------- #
def _style_header_row(ws, n_cols: int, row: int = 1) -> None:
    """Applica fill navy + font bianco grassetto + bordo alla riga header."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _set_widths(ws, widths: dict[str, int]) -> None:
    """Imposta la larghezza delle colonne. widths: {'A': 12, ...}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# --------------------------------------------------------------------------- #
# COSTRUZIONE DEI SINGOLI FOGLI
# --------------------------------------------------------------------------- #
def _write_transactions_sheet(wb: Workbook, rows: list[dict]) -> None:
    """Crea il foglio 'transactions' con header + righe fornite.

    Ogni riga è un dict con le chiavi di TX_HEADERS (notes opzionale).
    Applica number-format a date/quantity/price/fees per un'esperienza
    di compilazione pulita in Excel.
    """
    ws = wb.active
    ws.title = "transactions"
    ws.append(TX_HEADERS)
    _style_header_row(ws, len(TX_HEADERS))

    for r in rows:
        ws.append([r.get(h) for h in TX_HEADERS])

    # Number format per colonna (date=A, quantity=F, price=G, fees=I)
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=6).number_format = "#,##0.000000"
        ws.cell(row=i, column=7).number_format = "#,##0.0000"
        ws.cell(row=i, column=9).number_format = "#,##0.00"

    _set_widths(ws, {"A": 12, "B": 10, "C": 16, "D": 34, "E": 11,
                     "F": 12, "G": 12, "H": 10, "I": 8, "J": 22})
    ws.freeze_panes = "A2"


def _write_settings_sheet(wb: Workbook,
                          base_currency: str,
                          benchmark_ticker: str,
                          targets: dict[str, float]) -> None:
    """Crea il foglio 'settings' nel formato atteso da pf.load_settings.

    Layout (header=None lato lettura):
        base_currency     EUR
        benchmark_ticker  VWCE.DE
        <riga vuota>
        ticker            target_weight
        VWCE.DE           0.80
        VFEA.DE           0.20

    pf.load_settings cerca le chiavi note in col.A e il blocco allocation
    a partire dalla riga header 'ticker'/'...weight'.
    """
    ws = wb.create_sheet("settings")

    ws.append(["base_currency", base_currency])
    ws.append(["benchmark_ticker", benchmark_ticker])
    ws.append([])  # separatore

    header_row = ws.max_row + 1
    ws.append(["ticker", "target_weight"])
    _style_header_row(ws, 2, row=header_row)

    for tkr, w in targets.items():
        ws.append([tkr, w])
    for i in range(header_row + 1, ws.max_row + 1):
        ws.cell(row=i, column=2).number_format = "0.00"

    # Grassetto sulle chiavi di parametro
    ws.cell(row=1, column=1).font = KEY_FONT
    ws.cell(row=2, column=1).font = KEY_FONT
    _set_widths(ws, {"A": 20, "B": 16})


def _write_ter_sheet(wb: Workbook, tickers: list[str],
                     meta_map: dict | None = None) -> None:
    """Crea il foglio 'ter' (informativo, opzionale): ticker | ter_annual | note.

    meta_map permette di usare un'anagrafica diversa da quella del template
    (es. quella della demo, con ETF differenti).
    """
    meta_map = meta_map if meta_map is not None else _ETF_META
    ws = wb.create_sheet("ter")
    ws.append(["ticker", "ter_annual", "note"])
    _style_header_row(ws, 3)
    for t in tickers:
        meta = meta_map.get(t, {"ter": 0.0, "ter_note": ""})
        ws.append([t, meta["ter"], meta["ter_note"]])
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=2).number_format = "0.00%"
    _set_widths(ws, {"A": 12, "B": 14, "C": 46})


def _write_bollo_sheet(wb: Workbook, rows: list[dict] | None = None) -> None:
    """Crea il foglio 'bollo_charges' (opzionale): date | amount | notes.

    Se rows è None lascia solo header + una riga-nota esplicativa: l'utente
    inserirà qui gli addebiti reali del bollo che vede su Trade Republic.
    """
    ws = wb.create_sheet("bollo_charges")
    ws.append(["date", "amount", "notes"])
    _style_header_row(ws, 3)

    if rows:
        for r in rows:
            ws.append([r.get("date"), r.get("amount"), r.get("notes")])
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
            ws.cell(row=i, column=2).number_format = "#,##0.00"
    else:
        ws.append([None, None,
                   "Inserisci qui ogni addebito del bollo che vedi su TR "
                   "(facoltativo)"])
        ws["A2"].number_format = "yyyy-mm-dd"
        ws["B2"].number_format = "#,##0.00"
        ws["C2"].font = NOTE_FONT

    _set_widths(ws, {"A": 14, "B": 12, "C": 52})


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """Serializza il workbook in memoria e restituisce i bytes .xlsx."""
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# API PUBBLICA
# --------------------------------------------------------------------------- #
def build_template_workbook(with_examples: bool = True) -> bytes:
    """Genera il workbook TEMPLATE che l'utente scarica e compila.

    Contiene i 4 fogli con intestazioni corrette, settings/ter pre-compilati
    con i default VWCE/VFEA (80/20), e — se with_examples=True — due righe di
    esempio nel foglio transactions per mostrare il formato atteso.

    Returns:
        bytes del file .xlsx, pronti per st.download_button.
    """
    example_rows: list[dict] = []
    if with_examples:
        example_rows = [
            {"date": date(2025, 11, 10), "ticker": "VWCE.DE",
             "isin": "IE00BK5BQT80", "name": "Vanguard FTSE All-World (Acc)",
             "operation": "BUY", "quantity": 10.00, "price": 144.50,
             "currency": "EUR", "fees": 1.00, "notes": "ESEMPIO — sostituisci"},
            {"date": date(2025, 12, 8), "ticker": "VFEA.DE",
             "isin": "IE00BK5BR733",
             "name": "Vanguard FTSE Emerging Markets (Acc)",
             "operation": "BUY", "quantity": 20.00, "price": 70.20,
             "currency": "EUR", "fees": 1.00, "notes": "ESEMPIO — sostituisci"},
        ]

    wb = Workbook()
    _write_transactions_sheet(wb, example_rows)
    _write_settings_sheet(
        wb, base_currency="EUR", benchmark_ticker="VWCE.DE",
        targets={t: m["target"] for t, m in _ETF_META.items()},
    )
    _write_ter_sheet(wb, list(_ETF_META.keys()), _ETF_META)
    _write_bollo_sheet(wb, rows=None)
    return _workbook_to_bytes(wb)


def build_demo_workbook() -> bytes:
    """Genera il workbook DEMO: 3 ETF, ~7 anni di PAC con PREZZI EOD REALI.
 
    Serve la modalita' "prova senza compilare". Il portafoglio e' costruito
    con prezzi storici reali (scaricati da yfinance con build_demo_plan.py),
    cosi' Performance, Andamento, Rischio e Vs benchmark mostrano una storia
    autentica e non numeri inventati.
 
    Strategia rappresentata:
    - VWCE.DE (All-World): PAC mensile 500 € il 6 del mese, savings plan (fee 0);
    - SXR8.DE (S&P 500): 1500 €/anno il 15/06, acquisto manuale (fee 1 €);
    - EXSA.DE (Europe 600): 1500 €/anno il 15/12, acquisto manuale (fee 1 €).
    Le date sono i giorni di Borsa effettivi; le prime VWCE partono da
    agosto 2019 (l'ETF e' stato quotato a meta' 2019).
 
    Solo operazioni BUY, coerenti con la metodologia (cash-flow rebalancing).
    Il target 66/17/17 riflette il rapporto dei versamenti; con l'apprezzamento
    di VWCE i pesi di mercato lo superano, quindi la pagina Ribilanciamento
    produce un suggerimento concreto.
 
    Per rigenerare i prezzi (o cambiare periodo/regole) usa build_demo_plan.py
    e incolla qui il nuovo `plan`. Il tuple e' (data, ticker, qta, prezzo, fee).
 
    Returns:
        bytes del file .xlsx, da mettere in st.session_state["workbook_bytes"].
    """
    # (data, ticker, quantita', prezzo di carico, commissione) — prezzi EOD reali
    plan = [
        ("2019-06-17", "SXR8.DE", 5.909467, 253.8300, 1.0),
        ("2019-08-06", "VWCE.DE", 7.407407, 67.5000, 0.0),
        ("2019-09-06", "VWCE.DE", 6.993007, 71.5000, 0.0),
        ("2019-10-07", "VWCE.DE", 7.008691, 71.3400, 0.0),
        ("2019-11-06", "VWCE.DE", 6.717721, 74.4300, 0.0),
        ("2019-12-06", "VWCE.DE", 6.604147, 75.7100, 0.0),
        ("2019-12-16", "EXSA.DE", 36.456434, 41.1450, 1.0),
        ("2020-01-06", "VWCE.DE", 6.485084, 77.1000, 0.0),
        ("2020-02-06", "VWCE.DE", 6.242197, 80.1000, 0.0),
        ("2020-03-06", "VWCE.DE", 7.247427, 68.9900, 0.0),
        ("2020-04-06", "VWCE.DE", 7.946599, 62.9200, 0.0),
        ("2020-05-06", "VWCE.DE", 7.358352, 67.9500, 0.0),
        ("2020-06-08", "VWCE.DE", 6.773232, 73.8200, 0.0),
        ("2020-06-15", "SXR8.DE", 5.593258, 268.1800, 1.0),
        ("2020-07-06", "VWCE.DE", 6.774150, 73.8100, 0.0),
        ("2020-08-06", "VWCE.DE", 6.848377, 73.0100, 0.0),
        ("2020-09-07", "VWCE.DE", 6.660450, 75.0700, 0.0),
        ("2020-10-06", "VWCE.DE", 6.646284, 75.2300, 0.0),
        ("2020-11-06", "VWCE.DE", 6.495193, 76.9800, 0.0),
        ("2020-12-07", "VWCE.DE", 6.198859, 80.6600, 0.0),
        ("2020-12-15", "EXSA.DE", 38.659795, 38.8000, 1.0),
        ("2021-01-06", "VWCE.DE", 6.071646, 82.3500, 0.0),
        ("2021-02-08", "VWCE.DE", 5.771673, 86.6300, 0.0),
        ("2021-03-08", "VWCE.DE", 5.778343, 86.5300, 0.0),
        ("2021-04-06", "VWCE.DE", 5.518764, 90.6000, 0.0),
        ("2021-05-06", "VWCE.DE", 5.516937, 90.6300, 0.0),
        ("2021-06-07", "VWCE.DE", 5.459111, 91.5900, 0.0),
        ("2021-06-15", "SXR8.DE", 4.225828, 354.9600, 1.0),
        ("2021-07-06", "VWCE.DE", 5.257624, 95.1000, 0.0),
        ("2021-08-06", "VWCE.DE", 5.159426, 96.9100, 0.0),
        ("2021-09-06", "VWCE.DE", 5.057146, 98.8700, 0.0),
        ("2021-10-06", "VWCE.DE", 5.238345, 95.4500, 0.0),
        ("2021-11-08", "VWCE.DE", 4.879953, 102.4600, 0.0),
        ("2021-12-06", "VWCE.DE", 4.934860, 101.3200, 0.0),
        ("2021-12-15", "EXSA.DE", 32.303219, 46.4350, 1.0),
        ("2022-01-06", "VWCE.DE", 4.835590, 103.4000, 0.0),
        ("2022-02-07", "VWCE.DE", 5.064830, 98.7200, 0.0),
        ("2022-03-07", "VWCE.DE", 5.189414, 96.3500, 0.0),
        ("2022-04-06", "VWCE.DE", 4.945598, 101.1000, 0.0),
        ("2022-05-06", "VWCE.DE", 5.204538, 96.0700, 0.0),
        ("2022-06-06", "VWCE.DE", 5.166891, 96.7700, 0.0),
        ("2022-06-15", "SXR8.DE", 4.033885, 371.8500, 1.0),
        ("2022-07-06", "VWCE.DE", 5.383870, 92.8700, 0.0),
        ("2022-08-08", "VWCE.DE", 4.999000, 100.0200, 0.0),
        ("2022-09-06", "VWCE.DE", 5.159426, 96.9100, 0.0),
        ("2022-10-06", "VWCE.DE", 5.335040, 93.7200, 0.0),
        ("2022-11-07", "VWCE.DE", 5.357908, 93.3200, 0.0),
        ("2022-12-06", "VWCE.DE", 5.304477, 94.2600, 0.0),
        ("2022-12-15", "EXSA.DE", 35.323207, 42.4650, 1.0),
        ("2023-01-06", "VWCE.DE", 5.378658, 92.9600, 0.0),
        ("2023-02-06", "VWCE.DE", 5.157829, 96.9400, 0.0),
        ("2023-03-06", "VWCE.DE", 5.174912, 96.6200, 0.0),
        ("2023-04-06", "VWCE.DE", 5.269815, 94.8800, 0.0),
        ("2023-05-08", "VWCE.DE", 5.233958, 95.5300, 0.0),
        ("2023-06-06", "VWCE.DE", 4.980080, 100.4000, 0.0),
        ("2023-06-15", "SXR8.DE", 3.571769, 419.9600, 1.0),
        ("2023-07-06", "VWCE.DE", 5.019576, 99.6100, 0.0),
        ("2023-08-07", "VWCE.DE", 4.927079, 101.4800, 0.0),
        ("2023-09-06", "VWCE.DE", 4.866654, 102.7400, 0.0),
        ("2023-10-06", "VWCE.DE", 4.996003, 100.0800, 0.0),
        ("2023-11-06", "VWCE.DE", 4.973145, 100.5400, 0.0),
        ("2023-12-06", "VWCE.DE", 4.802151, 104.1200, 0.0),
        ("2023-12-15", "EXSA.DE", 31.823487, 47.1350, 1.0),
        ("2024-01-08", "VWCE.DE", 4.692192, 106.5600, 0.0),
        ("2024-02-06", "VWCE.DE", 4.461099, 112.0800, 0.0),
        ("2024-03-06", "VWCE.DE", 4.361479, 114.6400, 0.0),
        ("2024-04-08", "VWCE.DE", 4.270584, 117.0800, 0.0),
        ("2024-05-06", "VWCE.DE", 4.258218, 117.4200, 0.0),
        ("2024-06-06", "VWCE.DE", 4.169446, 119.9200, 0.0),
        ("2024-06-17", "SXR8.DE", 2.802691, 535.2000, 1.0),
        ("2024-07-08", "VWCE.DE", 4.042691, 123.6800, 0.0),
        ("2024-08-06", "VWCE.DE", 4.323016, 115.6600, 0.0),
        ("2024-09-06", "VWCE.DE", 4.207338, 118.8400, 0.0),
        ("2024-10-07", "VWCE.DE", 3.935768, 127.0400, 0.0),
        ("2024-11-06", "VWCE.DE", 3.833768, 130.4200, 0.0),
        ("2024-12-06", "VWCE.DE", 3.675390, 136.0400, 0.0),
        ("2024-12-16", "EXSA.DE", 29.434851, 50.9600, 1.0),
        ("2025-01-06", "VWCE.DE", 3.679717, 135.8800, 0.0),
        ("2025-02-06", "VWCE.DE", 3.600749, 138.8600, 0.0),
        ("2025-03-06", "VWCE.DE", 3.830830, 130.5200, 0.0),
        ("2025-04-07", "VWCE.DE", 4.443655, 112.5200, 0.0),
        ("2025-05-06", "VWCE.DE", 4.039425, 123.7800, 0.0),
        ("2025-06-06", "VWCE.DE", 3.840836, 130.1800, 0.0),
        ("2025-06-16", "SXR8.DE", 2.698133, 555.9400, 1.0),
        ("2025-07-07", "VWCE.DE", 3.825555, 130.7000, 0.0),
        ("2025-08-06", "VWCE.DE", 3.746441, 133.4600, 0.0),
        ("2025-09-08", "VWCE.DE", 3.664614, 136.4400, 0.0),
        ("2025-10-06", "VWCE.DE", 3.516669, 142.1800, 0.0),
        ("2025-11-06", "VWCE.DE", 3.487723, 143.3600, 0.0),
        ("2025-12-08", "VWCE.DE", 3.450656, 144.9000, 0.0),
        ("2025-12-15", "EXSA.DE", 26.041667, 57.6000, 1.0),
        ("2026-01-06", "VWCE.DE", 3.385240, 147.7000, 0.0),
        ("2026-02-06", "VWCE.DE", 3.382492, 147.8200, 0.0),
        ("2026-03-09", "VWCE.DE", 3.415301, 146.4000, 0.0),
        ("2026-04-07", "VWCE.DE", 3.474152, 143.9200, 0.0),
        ("2026-05-06", "VWCE.DE", 3.163356, 158.0600, 0.0),
        ("2026-06-08", "VWCE.DE", 3.094825, 161.5600, 0.0),
        ("2026-06-15", "SXR8.DE", 2.137970, 701.6000, 1.0),
    ]

    rows: list[dict] = []
    for iso, tkr, qty, px, fee in plan:
        y, mo, d = (int(x) for x in iso.split("-"))
        meta = _DEMO_META[tkr]
        rows.append({
            "date": date(y, mo, d), "ticker": tkr,
            "isin": meta["isin"], "name": meta["name"],
            "operation": "BUY", "quantity": qty, "price": px,
            "currency": "EUR", "fees": fee,
            "notes": f"PAC demo {y}-{mo:02d}",
        })

    wb = Workbook()
    _write_transactions_sheet(wb, rows)
    _write_settings_sheet(
        wb, base_currency="EUR", benchmark_ticker="VWCE.DE",
        targets={t: m["target"] for t, m in _DEMO_META.items()},
    )
    _write_ter_sheet(wb, list(_DEMO_META.keys()), _DEMO_META)
    _write_bollo_sheet(wb, rows=None)
    return _workbook_to_bytes(wb)


if __name__ == "__main__":
    # Smoke test manuale: scrive i due file su disco per ispezione a occhio.
    with open("transactions_template.xlsx", "wb") as f:
        f.write(build_template_workbook())
    with open("transactions_demo.xlsx", "wb") as f:
        f.write(build_demo_workbook())
    print("Scritti: transactions_template.xlsx, transactions_demo.xlsx")
