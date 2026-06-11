"""
add_costs_sheets.py — script una-tantum per aggiornare il file Excel.

Aggiunge due nuovi fogli al file transactions.xlsx esistente:
  - 'ter'           : TER per ticker (informativo)
  - 'bollo_charges' : addebiti reali del bollo da TR

Lascia intatti i fogli esistenti (transactions, settings).
Usage:
  python add_costs_sheets.py data/transactions.xlsx
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HDR_FILL = PatternFill('solid', start_color='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def add_costs_sheets(path: str):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f'File non trovato: {p}')

    wb = load_workbook(p)
    existing = wb.sheetnames
    added = []

    # ---- foglio 'ter' ----
    if 'ter' not in existing:
        ws = wb.create_sheet('ter')
        ws.append(['ticker', 'ter_annual', 'note'])
        ws.append(['VWCE.DE', 0.0022, '0.22% TER (Vanguard FTSE All-World)'])
        ws.append(['VFEA.DE', 0.0022, '0.22% TER (Vanguard FTSE Emerging Mkts)'])
        _style_header(ws, 3)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 45
        for r in (2, 3):
            ws.cell(row=r, column=2).number_format = '0.00%'
            ws.cell(row=r, column=2).font = Font(color='0000FF', name='Arial')
        added.append('ter')

    # ---- foglio 'bollo_charges' ----
    if 'bollo_charges' not in existing:
        ws = wb.create_sheet('bollo_charges')
        ws.append(['date', 'amount', 'notes'])
        ws.append([None, None,
                   "Inserisci qui ogni addebito del bollo che vedi su TR"])
        _style_header(ws, 3)
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws['A2'].number_format = 'yyyy-mm-dd'
        ws['B2'].number_format = '#,##0.00'
        added.append('bollo_charges')

    if added:
        wb.save(p)
        print(f'✓ Aggiornato {p}')
        print(f'  Fogli aggiunti: {", ".join(added)}')
    else:
        print(f'Nessuna modifica: i fogli "ter" e "bollo_charges" esistono già')
    print(f'  Fogli ora presenti: {wb.sheetnames}')


if __name__ == '__main__':
    target = sys.argv[1] if len(sys.argv) > 1 else 'data/transactions.xlsx'
    add_costs_sheets(target)
